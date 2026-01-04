import unittest
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook, Workbook

# cf_term_data.py から upsert_to_excel をインポート
from cf_term_data import upsert_to_excel

class TestUpsertToExcel(unittest.TestCase):

    def setUp(self):
        self.df = pd.DataFrame({
            'id': [1, 2, 3],
            'name': ['A', 'B', 'C'],
            'value': [10, 20, 30]
        })
        self.temp_dir = tempfile.mkdtemp()
        self.excel_file = os.path.join(self.temp_dir, 'test.xlsx')
        self.sheet_name = 'Sheet1'
        self.unique_index = 'id'

    def tearDown(self):
        if os.path.exists(self.excel_file):
            os.remove(self.excel_file)
        os.rmdir(self.temp_dir)

    def _read_excel(self):
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        headers = [cell.value for cell in ws[1]]
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        return headers, data

    def test_new_file_creation(self):
        """新規作成テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(headers, list(self.df.columns))
        self.assertEqual(data, self.df.values.tolist())

    def test_no_change_on_same_data(self):
        """再取得一致テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers1, data1 = self._read_excel()
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers2, data2 = self._read_excel()
        self.assertEqual(headers1, headers2)
        self.assertEqual(data1, data2)

    def test_update_changed_row(self):
        """変更行更新テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelを変更
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        ws.cell(row=3, column=3, value=999)  # value列の3行目を変更
        wb.save(self.excel_file)
        # 再実行
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(data[1][2], 20)  # 元に戻っている

    def test_add_new_row(self):
        """新規行追加テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        new_df = self.df.copy()
        new_df = pd.concat([new_df, pd.DataFrame({'id': [4], 'name': ['D'], 'value': [40]})], ignore_index=True)
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(len(data), 4)
        self.assertEqual(data[-1], [4, 'D', 40])

    def test_add_new_column(self):
        """新規列追加テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        new_df = self.df.copy()
        new_df['new_col'] = ['X', 'Y', 'Z']
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertIn('new_col', headers)
        self.assertEqual(data[0][3], 'X')
        self.assertEqual(data[1][3], 'Y')
        self.assertEqual(data[2][3], 'Z')

    def test_remove_row(self):
        """行削除対応テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelから行削除 (2行目を削除)
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        ws.delete_rows(3)  # ヘッダー+1行目なので、3行目削除でid=2の行
        wb.save(self.excel_file)
        # 再実行 (missing_rowsがあるので再書き込み)
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(len(data), 3)
        # 順序は保証されないかもしれないが、データが復活していることを確認
        ids = [row[0] for row in data]
        self.assertIn(1, ids)
        self.assertIn(2, ids)
        self.assertIn(3, ids)
        self.assertEqual(data[0], [1, 'A', 10])
        self.assertEqual(data[1], [3, 'C', 30])
        self.assertEqual(data[2], [2, 'B', 20])

    def test_missing_unique_column(self):
        """ユニークインデックス列欠損テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelからid列を削除
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        ws.delete_cols(1)  # id列（1列目）を削除
        wb.save(self.excel_file)
        # 再度実行 (unique_index列がないのでエラー)
        with self.assertRaises(ValueError) as cm:
            upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        self.assertIn("unique_index_label", str(cm.exception))

    def test_empty_dataframe_error(self):
        """空DataFrameエラーテスト"""
        empty_df = pd.DataFrame()
        with self.assertRaises(ValueError) as cm:
            upsert_to_excel(empty_df, self.sheet_name, self.excel_file, self.unique_index)
        self.assertIn("DataFrame must not be empty", str(cm.exception))

    def test_subset_columns_preserves_existing(self):
        """部分列DataFrameによる更新で既存列維持テスト"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        new_df = self.df[['id', 'name']]  # value列削除
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        # value列が残っていることを確認
        self.assertIn('value', headers)
        self.assertEqual(len(headers), 3)
        # データも残っているか確認
        value_idx = headers.index('value')
        self.assertEqual(data[0][value_idx], 10)
        self.assertEqual(data[1][value_idx], 20)
        self.assertEqual(data[2][value_idx], 30)

        new_df = self.df[['id', 'value']]  # name列削除
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        # name列が残っていることを確認
        self.assertIn('name', headers)
        self.assertEqual(len(headers), 3)
        # データも残っているか確認
        name_idx = headers.index('name')
        self.assertEqual(data[0][name_idx], 'A')
        self.assertEqual(data[1][name_idx], 'B')
        self.assertEqual(data[2][name_idx], 'C')

    def test_missing_unique_column_error(self):
        """ユニークインデックス列欠損エラーテスト"""
        # まず異なるヘッダーでシートを作成
        different_df = pd.DataFrame({'other_id': [1, 2, 3], 'data': ['X', 'Y', 'Z']})
        upsert_to_excel(different_df, self.sheet_name, self.excel_file, 'other_id')
        # unique_index_label が存在しない列でupsert
        with self.assertRaises(ValueError) as cm:
            upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        self.assertIn(f"unique_index_label '{self.unique_index}' column not found", str(cm.exception))

    def test_restore_missing_column(self):
        """列削除対応テスト（欠損列が復活）"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelから列削除 (value列を削除)
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        ws.delete_cols(3)  # value列（3列目）を削除
        wb.save(self.excel_file)
        # 再実行 (dfにはvalue列があるので復活するはず)
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertIn('value', headers)
        value_idx = headers.index('value')
        self.assertEqual(data[0][value_idx], 10)
        self.assertEqual(data[1][value_idx], 20)
        self.assertEqual(data[2][value_idx], 30)

    def test_column_order_change(self):
        """列順序変更テスト（既存順序維持）"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        new_df = self.df[['name', 'value', 'id']]  # 列順序変更
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        # 既存順序が維持される
        self.assertEqual(headers, ['id', 'name', 'value'])
        # データの内容が正しいか確認 (順序が変わっても値は正しい列に入るべき)
        id_idx = headers.index('id')
        name_idx = headers.index('name')
        val_idx = headers.index('value')
        self.assertEqual(data[0][id_idx], 1)
        self.assertEqual(data[0][name_idx], 'A')
        self.assertEqual(data[0][val_idx], 10)
        
        new_df = pd.DataFrame({'name': ['D'], 'value': [40], 'id': [4]})  # 新規行追加
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(len(data), 4)
        self.assertEqual(data[-1], [4, 'D', 40])
        

    def test_combination(self):
        """組み合わせテスト (新規行 + 新規列)"""
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        new_df = self.df.copy()
        new_df['new_col'] = ['X', 'Y', 'Z']
        new_df = pd.concat([new_df, pd.DataFrame({'id': [4], 'name': ['D'], 'value': [40], 'new_col': ['W']})], ignore_index=True)
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(len(headers), 4)
        self.assertEqual(len(data), 4)
        # 最後の行のデータを確認
        id_idx = headers.index('id')
        name_idx = headers.index('name')
        val_idx = headers.index('value')
        new_col_idx = headers.index('new_col')
        
        self.assertEqual(data[-1][id_idx], 4)
        self.assertEqual(data[-1][name_idx], 'D')
        self.assertEqual(data[-1][val_idx], 40)
        self.assertEqual(data[-1][new_col_idx], 'W')

    def test_empty_sheet(self):
        """空シートテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        wb.save(self.excel_file)
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        headers, data = self._read_excel()
        self.assertEqual(headers, list(self.df.columns))
        self.assertEqual(data, self.df.values.tolist())

    def test_sheet_not_exist(self):
        """シート不存在テスト"""
        # まずSheet1を作成
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # 次に存在しないSheet2でupsert
        new_sheet = 'Sheet2'
        upsert_to_excel(self.df, new_sheet, self.excel_file, self.unique_index, table_name='Table2')
        # Sheet2が存在することを確認
        wb = load_workbook(self.excel_file)
        self.assertIn(new_sheet, wb.sheetnames)
        # Sheet2のデータを確認
        ws = wb[new_sheet]
        headers = [cell.value for cell in ws[1]]
        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        self.assertEqual(headers, list(self.df.columns))
        self.assertEqual(data, self.df.values.tolist())

    def test_preserve_existing_custom_columns(self):
        """既存にしかないカスタム列が残るテスト"""
        # まず標準データを書き込み
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelにカスタム列を追加
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        ws.cell(row=1, column=4, value='custom_col')  # ヘッダーにカスタム列
        for r in range(2, 5):  # データ行にカスタム値を追加
            ws.cell(row=r, column=4, value=f'custom_{r-1}')
        wb.save(self.excel_file)
        # upsert（新規行追加）
        new_df = self.df.copy()
        new_df = pd.concat([new_df, pd.DataFrame({'id': [4], 'name': ['D'], 'value': [40]})], ignore_index=True)
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        # カスタム列が残っていることを確認
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        headers = [cell.value for cell in ws[1]]
        self.assertIn('custom_col', headers)
        custom_col_idx = headers.index('custom_col') + 1
        self.assertEqual(ws.cell(row=2, column=custom_col_idx).value, 'custom_1')
        self.assertEqual(ws.cell(row=3, column=custom_col_idx).value, 'custom_2')
        self.assertEqual(ws.cell(row=4, column=custom_col_idx).value, 'custom_3')
        # 新規行にもカスタム列がある（空）
        self.assertIsNone(ws.cell(row=5, column=custom_col_idx).value)

    def test_upsert_on_header_only_file(self):
        """ヘッダーだけ定義されたファイルに対するupsertテスト"""
        # ヘッダーだけのファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        for col_num, header in enumerate(self.df.columns, 1):
            ws.cell(row=1, column=col_num, value=header)
        wb.save(self.excel_file)
        # upsert_to_excel を実行
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # 結果を確認
        headers, data = self._read_excel()
        self.assertEqual(headers, list(self.df.columns))
        self.assertEqual(data, self.df.values.tolist())

    def test_upsert_with_distant_cell_write(self):
        """一度データをupsertした後、離れたセルに書き込み、再度書き込めるかを確認するテスト"""
        # まずデータをupsert
        upsert_to_excel(self.df, self.sheet_name, self.excel_file, self.unique_index)
        # Excelを開いて、離れたセルに書き込み
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        # データは行1-4（ヘッダー+3行）、列1-3
        # 離れたセル: 行10、列10 に書き込み
        ws.cell(row=10, column=10, value="distant_value")
        wb.save(self.excel_file)
        # 再度 upsert（データを変更して）
        new_df = self.df.copy()
        new_df.loc[0, 'value'] = 999  # 変更
        upsert_to_excel(new_df, self.sheet_name, self.excel_file, self.unique_index)
        # 結果を確認
        headers, data = self._read_excel()
        self.assertEqual(headers[:len(list(self.df.columns))], list(self.df.columns))
        self.assertGreater(len(headers), len(list(self.df.columns)))  # 列数は減っていない
        self.assertEqual(headers[len(list(self.df.columns))], None)

        # データが更新されているか
        self.assertEqual(data[0][2], 999)  # value が更新
        data_len = len(self.df)
        # data の長さが正しく、離れたセルを含まないことを確認
        self.assertGreater(len(data), data_len)  # データ行は3行
        # 各データ行が有効（Noneでない）ことを確認
        for row in data[:data_len]:
            self.assertIsNotNone(row[0])  # id が None でない
        self.assertEqual(data[data_len][0], None)  # 余分な行は None
        # 離れたセルの値が残っているか（wsに直接アクセス）
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        self.assertEqual(ws.cell(row=10, column=10).value, "distant_value")

if __name__ == '__main__':
    unittest.main()
