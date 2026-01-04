#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import logging
import datetime
import pandas as pd
import dateutil.parser
from datetime import timedelta, datetime
import sqlite3
from contextlib import closing
from time import sleep
from random import uniform
from tqdm import tqdm
import os
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

logger = logging.getLogger(__name__)

formatter = '%(levelname)s : %(asctime)s : %(message)s'
logging.basicConfig(level=logging.INFO, format=formatter)

# move HTTP request helpers to separate module
from moneyforward_api import *

# move shared utilities to separate module
from moneyforward_utils import traverse, get_categories_form_session


def is_range_overlapping(range1, range2):
    """
    Check if two ranges overlap.
    range1 and range2 are tuples (min_col, min_row, max_col, max_row)
    """
    min_col1, min_row1, max_col1, max_row1 = range1
    min_col2, min_row2, max_col2, max_row2 = range2
    return not (max_col1 < min_col2 or max_col2 < min_col1 or max_row1 < min_row2 or max_row2 < min_row1)


def save_workbook(wb, excel_file):
    """
    Save the workbook with retry on PermissionError.
    
    Args:
        wb: openpyxl.Workbook
        excel_file: str
    """
    while True:
        try:
            wb.save(excel_file)
            break
        except PermissionError:
            print(f"PermissionError: {excel_file} が開いている可能性があります。Excelファイルを閉じてEnterキーを押してください。")
            input()


def has_overlapped_range_table(ws, new_range: tuple, table_name: str) -> bool:
    if not ws.tables:
        return False
    
    for other_name, other_table in ws.tables.items():
        if other_name == table_name:
            continue
        other_range = range_boundaries(other_table)
        if is_range_overlapping(new_range, other_range):
            warnings.warn(f"Range overlap detected with table '{other_name}', skipping table operation for '{table_name}'.")
            return True
    return False

def add_new_table(ws, table_name: str, new_max_col: int, new_max_row: int):
    """
    Add a new Excel table to the worksheet.
    
    Args:
        ws: openpyxl の Worksheet
        table_name: str
        new_max_col: int
        new_max_row: int
    """
    if new_max_col <= 1 or new_max_row <= 1:
        warnings.warn(f"Cannot create table '{table_name}' with insufficient size: cols={new_max_col}, rows={new_max_row}.")
        return

    table_ref = f"A1:{get_column_letter(new_max_col)}{new_max_row}"
    new_range = (1, 1, new_max_col, new_max_row)
    # 重複チェック
    if has_overlapped_range_table(ws, new_range, table_name):
        return
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

def manage_table(ws, table_name: str, new_max_col: int, new_max_row: int):
    """
    Manage Excel table: create or update based on table_name.
    
    Args:
        ws: openpyxl.Worksheet
        table_name: str
        new_max_col: int(新規作成のみ使用。更新時は行のみ反映し、列は使用しない。Excelファイルが壊れるため)
        new_max_row: int
    """
    if ws.tables:
        # 引数のテーブル名が存在するか確認
        if table_name in ws.tables:
            table = ws.tables[table_name]
            min_col, min_row, max_col, _ = range_boundaries(table.ref)
            # 既存テーブルの範囲の、行のみ更新
            new_range = (min_col, min_row, max_col, new_max_row)
            # 重複チェック
            if has_overlapped_range_table(ws, new_range, table_name):
                return
            if not (min_col == 1 and min_row == 1):
                # A1を含まない場合、警告
                warnings.warn(f"Table '{table_name}' does not include A1, but expanding its range.")
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{new_max_row}"
            table.ref = new_ref
        else:
            # 引数のテーブル名が存在しない：A1を含むテーブルを探す
            a1_table = None
            for t in ws.tables.values():
                min_col, min_row, _, _ = range_boundaries(t.ref)
                if min_col == 1 and min_row == 1:
                    a1_table = t
                    break
            if a1_table:
                # A1を含むテーブルがある：そのテーブルを広げる、警告
                warnings.warn(f"Table '{table_name}' not found, using A1-containing table '{a1_table.displayName}' and expanding its range.")
                manage_table(ws, a1_table.displayName, new_max_col, new_max_row)
            else:
                # A1を含むテーブルがない：新規作成
                warnings.warn(f"Table '{table_name}' not found and no A1-containing table, creating new table '{table_name}'.")
                add_new_table(ws, table_name, new_max_col, new_max_row)
    else:
        # テーブルが存在しない：新規作成
        add_new_table(ws, table_name, new_max_col, new_max_row)


def upsert(frame, name: str, unique_index_label, con):
    pandas_sql = pd.io.sql.pandasSQL_builder(con)
    
    if isinstance(frame, pd.Series):
        frame = frame.to_frame()
    elif not isinstance(frame, pd.DataFrame):
        raise NotImplementedError(
            "'frame' argument should be either a Series or a DataFrame"
        )
    
    # table = pd.io.sql.SQLiteTable(name, pandas_sql, frame=frame, index=False, if_exists='append')
    table = pd.io.sql.SQLiteTable(name, pandas_sql, frame=frame, index=False, if_exists='append', keys=unique_index_label)
    table.create()
    # pandas_sql.execute('CREATE UNIQUE INDEX IF NOT EXISTS "{0}_{1}" ON "{0}" ("{1}");'.format(table.name, unique_index_label))
    
    def _execute_insert(self, conn, keys, data_iter):
        wld = "?"  # wildcard char
        escape = pd.io.sql._get_valid_sqlite_name

        bracketed_names = [escape(str(column)) for column in keys]
        col_names = ",".join(bracketed_names)
        wildcards = ",".join([wld] * len(bracketed_names))
        insert_statement = (
            f"INSERT OR REPLACE INTO {escape(self.name)} ({col_names}) VALUES ({wildcards})"
        )
        data_list = list(data_iter)
        conn.executemany(insert_statement, data_list)
    
    table.insert(method=_execute_insert)


def read_existing_data_from_sheet(ws, unique_index_label, sheet_name):
    """
    ワークシートから既存データを読み込み、headers と existing_df を返す。
    
    引数:
        ws: openpyxl.Worksheet オブジェクト。
        unique_index_label (str): ユニークインデックス列名。
        sheet_name (str): シート名（エラーメッセージ用）。
    
    戻り値:
        tuple: (existing_df, headers)
            existing_df: 既存データのDataFrame（'excel_row'列を含む）。
            headers: ヘッダーのリスト。
    
    例外:
        ValueError: 既存シートにunique_index_label列がない場合（headersが存在する場合のみ）。
    """
    existing_data = []
    row_numbers = []
    headers = []
    header_len = 0
    unique_idx = None
    existing_df = pd.DataFrame()
    existing_df['excel_row'] = []
    
    header_rows = list(ws.iter_rows(values_only=True, min_row=1, max_row=1))
    if not header_rows:
        return existing_df, headers

    # ヘッダー行を処理
    header_row = header_rows[0]
    for i, h in enumerate(header_row):
        if h is None:
            break
    
    header_len = i
    headers = header_row[:header_len]

    if not headers:
        return existing_df, headers

    # headersが存在する場合のみunique_index_label の存在確認
    if unique_index_label not in headers:
        raise ValueError(f"unique_index_label '{unique_index_label}' column not found in existing sheet '{sheet_name}': {headers}")
    unique_idx = headers.index(unique_index_label)
    
    # データ行を処理 (2行目から)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True, min_row=2), 2):
        val = row[unique_idx]
        if pd.isna(val) or val == '':
            break  # 無効な値なので、それ以降は読み込まない
        existing_data.append(row[:header_len])
        row_numbers.append(row_idx)
    
    if existing_data:
        existing_df = pd.DataFrame(existing_data, columns=headers)
        existing_df['excel_row'] = row_numbers
    
    return existing_df, headers



def load_excel_sheet(excel_file, sheet_name, unique_index_label):
    """
    Excelファイルとシートを読み込み、既存データを準備する。
    
    引数:
        excel_file (str): Excelファイルのパス。
        sheet_name (str): シート名。
        unique_index_label (str): ユニークインデックス列名。
    
    戻り値:
        tuple: (wb, ws, existing_df, headers)
            wb: openpyxl.Workbook オブジェクト。
            ws: ワークシートオブジェクト。
            existing_df: 既存データのDataFrame（'excel_row'列を含む）。
            headers: ヘッダーのリスト。
    
    例外:
        ValueError: 既存シートにunique_index_label列がない場合（headersが存在する場合のみ）。
    """
    if not os.path.exists(excel_file):
        # ファイルとシートを新規作成
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    elif sheet_name not in (wb := load_workbook(excel_file)).sheetnames:    
        # 既存のファイルを開き、シートが存在しない場合、シートのみ新規作成
        ws = wb.create_sheet(sheet_name)
    else:
        # シートが存在する場合
        ws = wb[sheet_name]
    
    existing_df, headers = read_existing_data_from_sheet(ws, unique_index_label, sheet_name)
    
    return wb, ws, existing_df, headers


def upsert_to_excel(df, sheet_name, excel_file, unique_index_label, table_name="DataTable"):
    """
    ユニークインデックスを用いて差分更新を行い、DataFrameをExcelシートにアップサートします。


    この関数はExcelファイルに対してアップサート操作を行います。
    ファイルやシートが存在しない場合は新規作成します。
    既存データがある場合は、unique_index_label列を使って一致する行に差分を反映し、
    新規データ（列・行）は既存の書式やカスタム列・行を維持したまま末尾に追加します。
    既存データにしかない列や行はそのまま残ります。

    引数:
        df (pd.DataFrame): アップサートするDataFrame。空であってはなりません。
        sheet_name (str): Excelシート名。
        excel_file (str): Excelファイルのパス。
        unique_index_label (str): 更新時に使うユニークインデックスとなる列名。dfおよび既存シート（存在する場合）に必須。
        table_name (str): テーブル名。デフォルトは"DataTable"。

    例外:
        ValueError: dfが空、unique_index_labelが空、または既存シートにunique_index_label列がない場合。
        PermissionError: ファイルがロックされている場合（ユーザー入力によるリトライ処理あり）。

    動作:
        - 新規ファイル/シート: dfを書き込み作成。
        - 既存シートでスキーマ一致: 差分更新・追加。
        - スキーマ不一致（列・行）: 差分更新（新規列追加、既存列維持）。
        - 空df: エラー。
        - 既存シートにunique_index_label列がない: エラー。
    """
    if df.empty:
        raise ValueError("DataFrame must not be empty")
    if not unique_index_label:
        raise ValueError("unique_index_label must be provided and not empty")
    
    wb, ws, existing_df, headers = load_excel_sheet(excel_file, sheet_name, unique_index_label)
    
    # 1. 列の同期
    # dfにある列がheadersになければ追加
    current_headers = list(headers)
    col_map = {name: i+1 for i, name in enumerate(current_headers)}
    
    for col in df.columns:
        if col not in col_map:
            # 新規列追加
            new_col_idx = len(current_headers) + 1
            ws.cell(row=1, column=new_col_idx, value=col)
            col_map[col] = new_col_idx
            current_headers.append(col)

    # 2. 行の同期
    # unique_index_label はチェック済み
    df_indexed = df.set_index(unique_index_label)

    if not existing_df.empty and unique_index_label in existing_df.columns:
        existing_df_indexed = existing_df.set_index(unique_index_label)
        existing_ids = set(existing_df_indexed.index)
    else:
        existing_df_indexed = pd.DataFrame()
        existing_ids = set()

    new_ids = set(df_indexed.index)
    
    # 2-1. 新規行の追加
    rows_to_add = new_ids - existing_ids
    if rows_to_add:
        start_row = 1 + len(existing_df) + 1  # ヘッダー行 + 既存データ行 の 次の行
        for idx in rows_to_add:
            row_data = df_indexed.loc[idx]
            # unique_index_label の値も書き込む
            if unique_index_label in col_map:
                 ws.cell(row=start_row, column=col_map[unique_index_label], value=idx)
            
            for col, val in row_data.items():
                if col in col_map:
                    ws.cell(row=start_row, column=col_map[col], value=val)
            start_row += 1

    # データのある最大行を計算 (ヘッダー + 既存データ行 + 新規追加行)
    data_max_row = 1 + len(existing_df) + len(rows_to_add)

    # 2-2. 既存行の更新
    common_ids = new_ids & existing_ids
    if common_ids:
        for idx in common_ids:
            excel_row_idx = existing_df_indexed.loc[idx, 'excel_row']
            if isinstance(excel_row_idx, pd.Series):
                excel_row_idx = excel_row_idx.iloc[0]
            
            row_data = df_indexed.loc[idx]
            for col, val in row_data.items():
                if col in col_map:
                    should_write = True
                    if col in existing_df_indexed.columns:
                        existing_val = existing_df_indexed.loc[idx, col]
                        # pandas の比較は NaN の扱いに注意
                        if pd.isna(val) and pd.isna(existing_val):
                            should_write = False
                        elif val == existing_val:
                            should_write = False
                    
                    if should_write:
                        ws.cell(row=excel_row_idx, column=col_map[col], value=val)
    
    # テーブル処理前に保存
    save_workbook(wb, excel_file)
    
    # テーブル処理
    manage_table(ws, table_name, len(current_headers), data_max_row)
    
    # 保存
    save_workbook(wb, excel_file)


def get_account_summaries_list(account_summaries, args):
    def ext(output_list, account):
        list_key1 = 'sub_accounts'
        list_key2 = 'user_asset_det_summaries'
        
        account_ref = {}
        traverse(account_ref, '', account, (list_key1, list_key2,))
        
        if list_key1 not in account or not account[list_key1]:
            output_list.append(account_ref)
            return
        
        for sub_accounts in account[list_key1]:
            account_ref1 = account_ref.copy()
            traverse(account_ref1, list_key1, sub_accounts, (list_key2,))
            
            if list_key2 not in sub_accounts or not sub_accounts[list_key2]:
                output_list.append(account_ref1)
                continue
            
            for user_asset_det_summaries in sub_accounts[list_key2]:
                account_ref2 = account_ref1.copy()
                traverse(account_ref2, list_key1 + "." + list_key2, user_asset_det_summaries, ())
                output_list.append(account_ref2)
    
    accounts = []
    for account in account_summaries['accounts']:
        ext(accounts, account)
    df = pd.DataFrame(accounts)
    
    if getattr(args, 'service_category_id', None):
        df = df[df['service_category_id'] == args.service_category_id]
    
    if getattr(args, 'name', None):
        df = df[df['name'].str.contains(args.name)]
    
    if getattr(args, 'sub_type', None):
        df = df[df['sub_accounts.sub_type'].str.contains(args.sub_type)]
    
    return df


def get_term_data_list(cf_term_data_by_sub_account, s=None, large=None, middle=None):
    user_asset_acts = []
    for e in cf_term_data_by_sub_account['user_asset_acts']:
        other = [k for k in e.keys() if k != 'user_asset_act']
        if other:
            print("other", other)
        
        user_asset_act_ref = {}
        traverse(user_asset_act_ref, '', e['user_asset_act'])
        user_asset_acts.append(user_asset_act_ref)
    
    if s:
        large, middle = get_categories_form_session(s)
    
    if large and middle:
        for i in range(len(user_asset_acts)):
            user_asset_acts[i]['large_category'] = large[user_asset_acts[i]['large_category_id']]
            user_asset_acts[i]['middle_category'] = middle[user_asset_acts[i]['middle_category_id']]
    
    for i, e in enumerate(user_asset_acts):
        user_asset_acts[i]['date'] = datetime.fromisoformat(e['recognized_at']).strftime("%y/%m/%d")
        user_asset_acts[i]['year'] = datetime.fromisoformat(e['recognized_at']).strftime("CY%y")
        user_asset_acts[i]['month'] = datetime.fromisoformat(e['recognized_at']).strftime("%y'%m")
    
    return pd.DataFrame(user_asset_acts)


def xrange(start, stop, step):
    while start <= stop:
        yield start
        start += step

def request_term_data(s, args):
    large, middle = get_categories_form_session(s)
    df = get_account_summaries_list(request_account_summaries(s), args)
    
    sub_account_id_hash_list = df['sub_accounts.sub_account_id_hash'].unique()
    
    def gen_prams(sub_account_id_hash_list, args):
        for sub_account_id_hash in tqdm(sub_account_id_hash_list, desc='sub_account'):
            params = dict(sub_account_id_hash=sub_account_id_hash,
                          date_from=args.date_from,
                          date_to=args.date_to)
            if not args.date_from or not args.date_to:
                yield params
                continue
            
            date_from_list = list(xrange(args.date_from, args.date_to, timedelta(days=365)))
            for date_from in tqdm(date_from_list, leave=False, desc='date'):
                params['date_from'] = date_from
                params['date_to'] = min(date_from + timedelta(days=364), args.date_to)
                yield params
    
    term_data_list = []
    for params in gen_prams(sub_account_id_hash_list, args):
        cf_term_data = request_cf_term_data_by_sub_account(s, **params)
        df = get_term_data_list(cf_term_data, large=large, middle=middle)
        term_data_list.append(df)
        sleep(uniform(0.1, 1))
    
    return pd.concat(term_data_list)


def get_term_data(s, args):
    with change_default_group(s):
        term_data_list = request_term_data(s, args)
    
    if args.csv:
        if args.csv_header:
            if args.ignore_KeyError:
                for c in set(args.csv_header) - set(term_data_list.columns):
                    term_data_list[c] = None
            term_data_list = term_data_list[args.csv_header]
        term_data_list.to_csv(args.csv, encoding='utf-8-sig', index=False)
        return
    
    if args.sqlite:
        if args.sqlite_header:
            select_header, rename_header, dtypes_dict = parse_header(args.sqlite_header)
            for c in set(select_header) - set(term_data_list.columns):
                term_data_list[c] = None
            term_data_list = term_data_list[select_header]
            
            if rename_header:
                term_data_list = term_data_list.rename(columns=rename_header)
            
            # dtypesを適用
            term_data_list = term_data_list.astype(dtypes_dict)
            
        with closing(sqlite3.connect(args.sqlite)) as con:
            upsert(term_data_list, 'user_asset_act', 'id', con)
        return
    
    if args.excel:
        if args.excel_header:
            select_header, rename_header, dtypes_dict = parse_header(args.excel_header)
            for c in set(select_header) - set(term_data_list.columns):
                term_data_list[c] = None
            term_data_list = term_data_list[select_header]
            
            if rename_header:
                term_data_list = term_data_list.rename(columns=rename_header)
            
            # dtypesを適用
            term_data_list = term_data_list.astype(dtypes_dict)
            
        upsert_to_excel(term_data_list, 'user_asset_act', args.excel, 'id')
        return
    
    print(*term_data_list.columns.tolist())
    for index, row in term_data_list.iterrows():
        print(*row.tolist())


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--mf_cookies', default='mf_cookies.pkl')
    parser.add_argument('-d', '--debug', action='store_true')
    
    parser.add_argument('-f', '--date_from', type=dateutil.parser.parse)
    parser.add_argument('-t', '--date_to', type=dateutil.parser.parse)
    parser.add_argument('-C', '--service_category_id', type=int)
    parser.add_argument('-N', '--name')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--csv')
    group.add_argument('--sqlite')
    group.add_argument('--excel')
    parser.add_argument('--csv_header', nargs='+')
    sqlite_header = """id:str date year month account_id:str sub_account_id:str is_transfer is_income
                       orig_content=content orig_amount=amount currency jpyrate memo 
                       large_category_id middle_category_id large_category middle_category
                       is_target partner_account_id:str partner_sub_account_id:str partner_act_id:str
                       created_at recognized_at updated_at sub_account_id_hash transfer_type
                       account.account.service_id=service_id
                       account.account.service_category_id=service_category_id
                       account.account.disp_name=disp_name
                       account.account.service.service.service_name=service_name
                       sub_account.sub_account.sub_name=sub_name
                       sub_account.sub_account.sub_type=sub_type
                       sub_account.sub_account.sub_number=sub_number
                       partner_account.partner_account.service_id=partner_account_service_id
                       partner_account.partner_account.service_category_id=partner_account_service_category_id
                       partner_account.partner_account.disp_name=partner_account_disp_name
                       partner_account.partner_account.memo=partner_account_memo
                       partner_account.partner_account.display_name=partner_account_display_name
                       partner_sub_account.partner_sub_account.sub_name=partner_account_sub_name
                       partner_sub_account.partner_sub_account.sub_type=partner_account_sub_type
                       partner_sub_account.partner_sub_account.sub_number=partner_account_sub_number
                       partner_sub_account.partner_sub_account.service_category_id=partner_sub_account_service_category_id
                       partner_sub_account.partner_sub_account.is_dummy=partner_sub_account_is_dummy
                       partner_act.partner_act.orig_content=partner_act_content
                       partner_act.partner_act.orig_amount=partner_act_amount
                       partner_act.partner_act.currency=partner_act_currency
                       partner_act.partner_act.jpyrate=partner_act_jpyrate
                       partner_act.partner_act.memo=partner_act_memo
                       partner_act.partner_act.large_category_id=partner_act_large_category_id
                       partner_act.partner_act.middle_category_id=partner_act_middle_category_id
                       partner_act.partner_act.sub_account_id_hash=partner_act_sub_account_id_hash
                       partner_act.partner_act.partner_sub_account_id_hash=partner_act_partner_sub_account_id_hash
                       """.split()
    parser.add_argument('--sqlite_header', nargs='+', default=sqlite_header)
    parser.add_argument('--excel_header', nargs='+', default=sqlite_header)  # 同じデフォルトを使用
    parser.add_argument('-i', '--ignore_KeyError', action='store_true')
    
    args = parser.parse_args(argv)

    if args.debug:
        import http.client
        http.client.HTTPConnection.debuglevel = 2

    with session_from_cookie_file(args.mf_cookies) as s:
        get_term_data(s, args)


if __name__ == '__main__':
    main()